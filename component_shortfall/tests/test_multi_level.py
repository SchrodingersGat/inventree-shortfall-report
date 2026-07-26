"""Multi-level BOM regression tests for `calculate_shortfall`.

Fixture shape - a diamond BOM, so a single bottom-level component is reachable
via two independent paths with different quantity multipliers:

    TOP
    |-- MID_A (x2)
    |     |-- BOTTOM (x5)
    |-- MID_B (x3)
          |-- BOTTOM (x2)

A sales order for TOP therefore drives BOTTOM's requirement through *both*
branches, and the two contributions must be summed rather than overwritten.
"""

from decimal import Decimal

from build.models import Build
from common.models import DataOutput
from company.models import Company
from order.models import SalesOrder, SalesOrderLineItem
from part.models import BomItem, Part
from stock.models import StockItem, StockLocation

from InvenTree.unit_test import InvenTreeTestCase

from .. import shortfall


class DiamondBOMTestCase(InvenTreeTestCase):
    """Base test case which builds the diamond BOM fixture described above."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()

        cls.top = Part.objects.create(
            name='Top', assembly=True, component=False, salable=True
        )
        cls.mid_a = Part.objects.create(
            name='Mid A', assembly=True, component=True, salable=False
        )
        cls.mid_b = Part.objects.create(
            name='Mid B', assembly=True, component=True, salable=False
        )
        cls.bottom = Part.objects.create(
            name='Bottom', assembly=False, component=True, purchaseable=True
        )

        # Refresh - MPTT's tree_id fixup on save isn't reflected on these
        # in-memory instances, which trips the BOM 'recursive' check otherwise.
        for part in [cls.top, cls.mid_a, cls.mid_b, cls.bottom]:
            part.refresh_from_db()

        cls.item_top_a = BomItem.objects.create(
            part=cls.top, sub_part=cls.mid_a, quantity=2
        )
        cls.item_top_b = BomItem.objects.create(
            part=cls.top, sub_part=cls.mid_b, quantity=3
        )
        cls.item_a_bottom = BomItem.objects.create(
            part=cls.mid_a, sub_part=cls.bottom, quantity=5
        )
        cls.item_b_bottom = BomItem.objects.create(
            part=cls.mid_b, sub_part=cls.bottom, quantity=2
        )

        cls.customer = Company.objects.create(name='Customer', is_customer=True)

    def make_output(self):
        """Create a DataOutput instance to pass into calculate_shortfall."""
        return DataOutput.objects.create(
            total=0, progress=0, output_type='shortfall_report'
        )

    def create_sales_order(self, part, quantity):
        """Helper - create a sales order + single line item for the given part."""
        so = SalesOrder.objects.create(
            customer=self.customer,
            reference=f'SO-{SalesOrder.objects.count() + 1:04d}',
        )
        SalesOrderLineItem.objects.create(order=so, part=part, quantity=quantity)
        return so


class DiamondBOMQuantityTests(DiamondBOMTestCase):
    """Verify quantity multipliers compound correctly through both paths of the diamond."""

    def test_bottom_requirement_sums_both_paths(self):
        """BOTTOM's required quantity is the sum of both branch contributions."""
        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['required'], 10)
        self.assertEqual(requirements[self.mid_a.pk]['required'], 20)  # 10 * 2
        self.assertEqual(requirements[self.mid_b.pk]['required'], 30)  # 10 * 3

        # Bottom: (10 * 2 * 5) + (10 * 3 * 2) = 100 + 60 = 160
        self.assertEqual(requirements[self.bottom.pk]['required'], 160)
        self.assertEqual(requirements[self.bottom.pk]['shortfall'], 160)

    def test_intermediate_stock_reduces_downstream_requirement(self):
        """Stock held at MID_A only reduces BOTTOM's requirement via the MID_A branch."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(part=self.mid_a, quantity=5, location=location)

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        # MID_A: required 20, stock 5 -> shortfall 15 (only the deficit propagates)
        self.assertEqual(requirements[self.mid_a.pk]['shortfall'], 15)

        # Bottom: (15 * 5) [via MID_A's shortfall] + (30 * 2) [via MID_B, no stock] = 75 + 60 = 135
        self.assertEqual(requirements[self.bottom.pk]['required'], 135)

    def test_full_stock_at_intermediate_level_stops_that_branch(self):
        """Full coverage of MID_B's requirement via stock stops propagation down that branch only."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(part=self.mid_b, quantity=100, location=location)

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.mid_b.pk]['shortfall'], 0)

        # Bottom only receives the MID_A contribution: 10 * 2 * 5 = 100
        self.assertEqual(requirements[self.bottom.pk]['required'], 100)

    def test_on_order_and_in_production_offset_requirement(self):
        """On-order and in-production quantities for a part offset its own shortfall."""
        # Give MID_A an open purchase order and an active build, to populate
        # on_order / quantity_being_built without requiring stock.
        supplier = Company.objects.create(name='Supplier', is_supplier=True)
        from company.models import SupplierPart
        from order.models import PurchaseOrder, PurchaseOrderLineItem

        supplier_part = SupplierPart.objects.create(
            part=self.mid_a, supplier=supplier, SKU='SKU-MIDA'
        )
        po = PurchaseOrder.objects.create(supplier=supplier, reference='PO-0001')
        PurchaseOrderLineItem.objects.create(
            order=po, part=supplier_part, quantity=8
        )

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        # MID_A: required 20, on_order 8 -> shortfall 12
        self.assertEqual(requirements[self.mid_a.pk]['on_order'], 8)
        self.assertEqual(requirements[self.mid_a.pk]['shortfall'], 12)


class MultipleDemandSourcesTests(DiamondBOMTestCase):
    """Verify requirements from multiple independent demand sources are aggregated correctly."""

    def test_two_sales_orders_for_same_top_part_aggregate(self):
        """Two separate sales orders for TOP aggregate their requirements at every level."""
        self.create_sales_order(self.top, 10)
        self.create_sales_order(self.top, 5)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['required'], 15)
        self.assertEqual(requirements[self.mid_a.pk]['required'], 30)
        # Bottom: (15 * 2 * 5) + (15 * 3 * 2) = 150 + 90 = 240
        self.assertEqual(requirements[self.bottom.pk]['required'], 240)

    def test_direct_and_indirect_demand_for_same_part_aggregate(self):
        """A part required both directly (its own sales order) and indirectly (via a parent BOM) aggregates correctly."""
        # Direct sales order for the (normally sub-) component MID_A, as if it
        # were also independently salable/spare-part-orderable.
        self.create_sales_order(self.mid_a, 4)
        # Indirect demand for MID_A via TOP's BOM
        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        # MID_A: 4 (direct) + 20 (10 * 2 via TOP) = 24
        self.assertEqual(requirements[self.mid_a.pk]['required'], 24)

    def test_build_order_and_sales_order_demand_aggregate(self):
        """Demand for MID_A from an active TOP build order and a TOP sales order are combined.

        Note: an active build order for TOP contributes demand in *two* ways:
        - Directly, via its BuildLines - MID_A/MID_B are needed to complete the
          5 TOPs currently in production.
        - Indirectly, TOP's own `quantity_being_built` (5) offsets TOP's *own*
          shortfall (required 10 - in_production 5 = shortfall 5), so only the
          remaining 5 TOPs (not all 10) propagate a *further* MID_A/MID_B need.

        Overall this is consistent, not double-counted: 5 TOPs already in
        production (10 MID_A via BuildLines) + 5 more TOPs still needed (10 MID_A
        via the sales-order shortfall) = 10 TOPs' worth of MID_A = 20.
        """
        self.create_sales_order(self.top, 10)
        Build.objects.create(part=self.top, quantity=5, reference='BO-0001')

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['required'], 10)
        self.assertEqual(requirements[self.top.pk]['shortfall'], 5)
        self.assertEqual(requirements[self.mid_a.pk]['required'], 20)
        self.assertEqual(requirements[self.mid_b.pk]['required'], 30)


class ShortfallReportOutputTests(DiamondBOMTestCase):
    """Verify the generated XLSX report content respects filtering options."""

    def load_report_rows(self, output):
        """Load the generated XLSX file and return its data rows (excluding header)."""
        from openpyxl import load_workbook

        output.refresh_from_db()
        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return rows

    def test_hide_no_shortfall_omits_covered_parts(self):
        """Parts with zero shortfall are omitted from the report when hide_no_shortfall=True."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(part=self.mid_b, quantity=1000, location=location)

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        shortfall.calculate_shortfall(output.pk, hide_no_shortfall=True)

        rows = self.load_report_rows(output)
        names = {row[0] for row in rows}

        self.assertIn('Top', names)
        self.assertIn('Mid A', names)
        self.assertNotIn('Mid B', names)

    def test_category_filter_limits_report_rows(self):
        """Only parts within the selected category (or its descendants) appear in the report."""
        from part.models import PartCategory

        category = PartCategory.objects.create(name='Bottom Category')
        self.bottom.category = category
        self.bottom.save()

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False, category_id=category.pk
        )

        rows = self.load_report_rows(output)
        names = {row[0] for row in rows}

        self.assertEqual(names, {'Bottom'})

    def test_supplier_data_included_when_requested(self):
        """Supplier names are appended as an extra column when include_supplier_data=True."""
        supplier = Company.objects.create(name='Acme Supply', is_supplier=True)
        from company.models import SupplierPart

        SupplierPart.objects.create(
            part=self.bottom, supplier=supplier, SKU='SKU-BOTTOM'
        )

        self.create_sales_order(self.top, 10)

        output = self.make_output()
        shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False, include_supplier_data=True
        )

        output.refresh_from_db()
        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn('Suppliers', header)

        rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
        supplier_col = header.index('Suppliers')
        self.assertEqual(rows['Bottom'][supplier_col], 'Acme Supply')
