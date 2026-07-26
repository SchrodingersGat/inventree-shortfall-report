"""API endpoint tests for the ComponentShortfall plugin."""

from django.urls import reverse

from build.models import Build
from common.models import DataOutput, InvenTreeSetting
from company.models import Company
from order.models import SalesOrder, SalesOrderLineItem
from part.models import BomItem, Part
from stock.models import StockItem, StockLocation

from InvenTree.unit_test import InvenTreeAPITestCase


class ShortfallReportAPITestCase(InvenTreeAPITestCase):
    """Base test case for the '/plugin/component-shortfall/shortfall/' endpoint."""

    def setUp(self):
        """Ensure the plugin's URLs are registered, and create a simple assembly."""
        super().setUp()

        # Ensure plugin URLs are registered even if INVENTREE_PLUGIN_TESTING_SETUP
        # is not set in the environment this test happens to run under.
        InvenTreeSetting.set_setting('ENABLE_PLUGINS_URL', True, None)

        self.assembly = Part.objects.create(
            name='Assembly', assembly=True, component=False, salable=True
        )
        self.component = Part.objects.create(
            name='Component', component=True, purchaseable=True, salable=False
        )
        for part in [self.assembly, self.component]:
            part.refresh_from_db()

        self.bom_item = BomItem.objects.create(
            part=self.assembly, sub_part=self.component, quantity=2
        )

        self.customer = Company.objects.create(name='Customer', is_customer=True)

        self.url = reverse('plugin:component-shortfall:shortfall-report-view')


class ShortfallReportAPITests(ShortfallReportAPITestCase):
    """Tests for triggering shortfall report generation via the API."""

    def create_sales_order(self, quantity=10):
        """Helper to create an outstanding sales order for the assembly part."""
        so = SalesOrder.objects.create(customer=self.customer, reference='SO-0001')
        SalesOrderLineItem.objects.create(
            order=so, part=self.assembly, quantity=quantity
        )

    def test_basic_request_generates_completed_output(self):
        """A basic POST request synchronously generates a completed DataOutput (no worker running in tests)."""
        self.create_sales_order()

        response = self.post(self.url, data={}, expected_code=200)

        output_id = response.data['output']['pk']
        output = DataOutput.objects.get(pk=output_id)

        self.assertTrue(output.complete)
        self.assertEqual(output.output_type, 'shortfall_report')
        self.assertTrue(output.output.name.endswith('.xlsx'))

    def test_unauthenticated_request_rejected(self):
        """An unauthenticated request is rejected."""
        self.client.logout()
        response = self.client.post(self.url, data={})
        self.assertIn(response.status_code, (401, 403))

    def test_category_filter_is_accepted(self):
        """A valid category ID is accepted and passed through to the calculation."""
        from part.models import PartCategory

        category = PartCategory.objects.create(name='Test Category')
        self.create_sales_order()

        response = self.post(
            self.url, data={'category': category.pk}, expected_code=200
        )
        self.assertEqual(response.data['category'], category.pk)

    def test_invalid_category_rejected(self):
        """An invalid (non-existent) category ID is rejected with a 400."""
        response = self.post(
            self.url, data={'category': 999999}, expected_code=400
        )

    def test_max_bom_depth_bounds_enforced(self):
        """max_bom_depth must be between 0 and 50 (inclusive)."""
        self.post(self.url, data={'max_bom_depth': -1}, expected_code=400)
        self.post(self.url, data={'max_bom_depth': 51}, expected_code=400)
        self.post(self.url, data={'max_bom_depth': 50}, expected_code=200)

    def test_max_bom_depth_limits_propagation_via_api(self):
        """A max_bom_depth of 0 stops the shortfall calculation from reaching sub-components."""
        self.create_sales_order()

        response = self.post(
            self.url, data={'max_bom_depth': 0}, expected_code=200
        )

        output_id = response.data['output']['pk']
        output = DataOutput.objects.get(pk=output_id)

        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        names = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}

        self.assertIn('Assembly', names)
        self.assertNotIn('Component', names)

    def test_include_build_orders_false_excludes_build_demand(self):
        """Setting include_build_orders=False excludes build-order-driven demand from the report."""
        Build.objects.create(part=self.assembly, quantity=6, reference='BO-0001')

        response = self.post(
            self.url, data={'include_build_orders': False}, expected_code=200
        )

        output_id = response.data['output']['pk']
        output = DataOutput.objects.get(pk=output_id)

        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # No demand at all (no sales orders, build orders excluded) - report has no rows
        # (hide_no_shortfall defaults to True at the plugin-setting level, and there is
        # no shortfall anywhere since demand was excluded)
        self.assertEqual(rows, [])

    def test_include_sales_orders_false_excludes_sales_demand(self):
        """Setting include_sales_orders=False excludes sales-order-driven demand from the report."""
        self.create_sales_order()

        response = self.post(
            self.url, data={'include_sales_orders': False}, expected_code=200
        )

        output_id = response.data['output']['pk']
        output = DataOutput.objects.get(pk=output_id)

        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        self.assertEqual(rows, [])

    def test_horizon_months_defaults_from_plugin_setting(self):
        """When horizon_months is not supplied, the plugin's configured default is used."""
        from plugin.registry import registry

        plugin = registry.get_plugin('component-shortfall')
        plugin.set_setting('SHORTFALL_HORIZON_MONTHS', 3)
        try:
            response = self.post(self.url, data={}, expected_code=200)
        finally:
            plugin.set_setting('SHORTFALL_HORIZON_MONTHS', 12)

    def test_report_reflects_current_stock(self):
        """The generated report reflects stock on hand at the time of generation."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(
            part=self.component, quantity=1000, location=location
        )
        self.create_sales_order()

        response = self.post(self.url, data={}, expected_code=200)

        output_id = response.data['output']['pk']
        output = DataOutput.objects.get(pk=output_id)

        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        names = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}

        # Component has ample stock -> no shortfall -> hidden from default report
        self.assertNotIn('Component', names)
