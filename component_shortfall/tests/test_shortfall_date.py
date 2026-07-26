"""Regression tests for the shortfall-date calculation (Phase 3)."""

from datetime import timedelta
from decimal import Decimal

from build.models import Build
from common.models import DataOutput
from company.models import Company, SupplierPart
from order.models import (
    PurchaseOrder,
    PurchaseOrderLineItem,
    SalesOrder,
    SalesOrderLineItem,
)
from part.models import BomItem, Part
from stock.models import StockItem, StockLocation

from InvenTree.helpers import current_date
from InvenTree.unit_test import InvenTreeTestCase

from .. import shortfall

TODAY = current_date()
IN_5_DAYS = TODAY + timedelta(days=5)
IN_10_DAYS = TODAY + timedelta(days=10)
IN_20_DAYS = TODAY + timedelta(days=20)


class FindShortfallDateTests(InvenTreeTestCase):
    """Unit tests for the `find_shortfall_date` primitive."""

    def test_already_negative_returns_today(self):
        """A negative starting balance always resolves to today, regardless of dated events."""
        result = shortfall.find_shortfall_date(Decimal(-5), [(IN_20_DAYS, Decimal(100))])
        self.assertEqual(result, TODAY)

    def test_negative_starting_balance_with_no_events(self):
        """A negative starting balance resolves to today even with an empty event list."""
        result = shortfall.find_shortfall_date(Decimal(-1), [])
        self.assertEqual(result, TODAY)

    def test_never_goes_negative_returns_none(self):
        """A balance that stays non-negative throughout returns None."""
        result = shortfall.find_shortfall_date(
            Decimal(10), [(IN_5_DAYS, Decimal(-5)), (IN_10_DAYS, Decimal(5))]
        )
        self.assertIsNone(result)

    def test_single_event_pushes_negative(self):
        """The date of the event which first pushes the balance negative is returned."""
        result = shortfall.find_shortfall_date(Decimal(10), [(IN_5_DAYS, Decimal(-15))])
        self.assertEqual(result, IN_5_DAYS)

    def test_recovers_after_later_supply_still_reports_first_negative_date(self):
        """Once the balance goes negative, later incoming supply doesn't erase the shortfall date."""
        result = shortfall.find_shortfall_date(
            Decimal(0),
            [(IN_5_DAYS, Decimal(-10)), (IN_10_DAYS, Decimal(100))],
        )
        self.assertEqual(result, IN_5_DAYS)

    def test_events_out_of_input_order_are_still_processed_chronologically(self):
        """Events are sorted internally - input order must not affect the result."""
        result = shortfall.find_shortfall_date(
            Decimal(5),
            [(IN_10_DAYS, Decimal(-100)), (IN_5_DAYS, Decimal(-2))],
        )
        # At day 5: 5 - 2 = 3 (still positive). At day 10: 3 - 100 = negative.
        self.assertEqual(result, IN_10_DAYS)

    def test_same_day_events_are_netted_before_checking_negativity(self):
        """Multiple events on the same date are summed together, not applied one-at-a-time.

        A -10 and +15 on the same day nets to +5 (never negative) - if they were
        applied in an arbitrary one-at-a-time order, processing the -10 first
        would incorrectly report a shortfall on that date.
        """
        result = shortfall.find_shortfall_date(
            Decimal(0),
            [(IN_5_DAYS, Decimal(-10)), (IN_5_DAYS, Decimal(15))],
        )
        self.assertIsNone(result)

    def test_exact_zero_balance_is_not_a_shortfall(self):
        """A balance of exactly zero is not considered a shortfall."""
        result = shortfall.find_shortfall_date(Decimal(0), [(IN_5_DAYS, Decimal(0))])
        self.assertIsNone(result)


class ShortfallDateFixtureTestCase(InvenTreeTestCase):
    """Base fixture: a simple TOP -> MID -> BOTTOM chain for end-to-end date tests."""

    def setUp(self):
        super().setUp()

        self.top = Part.objects.create(
            name='DateTop', assembly=True, component=False, salable=True
        )
        self.mid = Part.objects.create(
            name='DateMid', assembly=True, component=True, purchaseable=False
        )
        self.bottom = Part.objects.create(
            name='DateBottom', component=True, purchaseable=True
        )
        for part in [self.top, self.mid, self.bottom]:
            part.refresh_from_db()

        BomItem.objects.create(part=self.top, sub_part=self.mid, quantity=1)
        BomItem.objects.create(part=self.mid, sub_part=self.bottom, quantity=1)

        self.customer = Company.objects.create(name='Customer', is_customer=True)
        self.supplier = Company.objects.create(name='Supplier', is_supplier=True)

    def make_output(self):
        return DataOutput.objects.create(
            total=0, progress=0, output_type='shortfall_report'
        )

    def create_sales_order(self, part, quantity, target_date=None):
        so = SalesOrder.objects.create(
            customer=self.customer,
            reference=f'SO-{SalesOrder.objects.count() + 1:04d}',
            target_date=target_date,
        )
        SalesOrderLineItem.objects.create(order=so, part=part, quantity=quantity)
        return so

    def create_purchase_order(self, part, quantity, target_date=None):
        supplier_part, _ = SupplierPart.objects.get_or_create(
            part=part, supplier=self.supplier, SKU=f'SKU-{part.pk}'
        )
        po = PurchaseOrder.objects.create(
            supplier=self.supplier,
            reference=f'PO-{PurchaseOrder.objects.count() + 1:04d}',
            target_date=target_date,
        )
        PurchaseOrderLineItem.objects.create(
            order=po, part=supplier_part, quantity=quantity
        )
        return po


class EdgeCaseTests(ShortfallDateFixtureTestCase):
    """Tests for the two edge cases called out in the original request."""

    def test_insufficient_stock_now_gives_today(self):
        """A part with zero stock and an outstanding (undated) order is short as of today."""
        self.create_sales_order(self.top, 10)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['shortfall_date'], TODAY)

    def test_runs_out_before_incoming_arrives(self):
        """A part with equal total demand/supply can still show a shortfall date if supply arrives too late.

        Demand: 10 units due in 5 days. Supply: 10 units incoming, due in 10 days
        (after the demand). The scalar `shortfall` is 0 (10 required - 10 on
        order), but the part genuinely runs dry between day 5 and day 10.
        """
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)
        self.create_purchase_order(self.top, 10, target_date=IN_10_DAYS)

        # Ensure top is purchaseable too, so on_order/dated-supply applies to it directly
        self.top.purchaseable = True
        self.top.save()

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['shortfall'], 0)
        self.assertEqual(requirements[self.top.pk]['shortfall_date'], IN_5_DAYS)

    def test_sufficient_stock_and_timing_gives_no_shortfall_date(self):
        """A part with ample stock and no risky timing gap has no shortfall date at all."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(part=self.top, quantity=100, location=location)

        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertIsNone(requirements[self.top.pk]['shortfall_date'])

    def test_supply_arriving_before_demand_avoids_shortfall_date(self):
        """If incoming supply arrives *before* the demand needs it, there's no shortfall date."""
        self.create_sales_order(self.top, 10, target_date=IN_10_DAYS)
        self.create_purchase_order(self.top, 10, target_date=IN_5_DAYS)
        self.top.purchaseable = True
        self.top.save()

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertIsNone(requirements[self.top.pk]['shortfall_date'])


class CascadeDateTests(ShortfallDateFixtureTestCase):
    """Tests for propagating shortfall dates down the BOM cascade."""

    def test_sub_part_inherits_parent_shortfall_date(self):
        """MID's cascaded shortfall date matches TOP's own resolved shortfall date."""
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.top.pk]['shortfall_date'], IN_5_DAYS)
        self.assertEqual(requirements[self.mid.pk]['shortfall_date'], IN_5_DAYS)
        self.assertEqual(requirements[self.bottom.pk]['shortfall_date'], IN_5_DAYS)

    def test_earliest_of_multiple_top_level_orders_wins(self):
        """When two independent top-level orders both cascade to the same sub-part, the earlier date wins."""
        other_top = Part.objects.create(
            name='DateOtherTop', assembly=True, component=False, salable=True
        )
        other_top.refresh_from_db()
        BomItem.objects.create(part=other_top, sub_part=self.mid, quantity=1)

        # TOP is short sooner (5 days) than OTHER_TOP (20 days)
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)
        self.create_sales_order(other_top, 10, target_date=IN_20_DAYS)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertEqual(requirements[self.mid.pk]['shortfall_date'], IN_5_DAYS)

    def test_parent_with_no_shortfall_date_contributes_no_cascade_demand(self):
        """A parent that never goes short contributes no cascaded demand event to its children."""
        location = StockLocation.objects.create(name='Location')
        # TOP has ample stock - its own shortfall_date is None
        StockItem.objects.create(part=self.top, quantity=1000, location=location)
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        self.assertIsNone(requirements[self.top.pk]['shortfall_date'])
        # MID was never visited at all, since TOP's *scalar* shortfall is 0
        # (ample stock) - the existing recursion-trigger boundary is unchanged
        # by the date feature.
        self.assertNotIn(self.mid.pk, requirements)

    def test_diamond_multi_path_uses_earliest_resolved_parent_date(self):
        """A part reachable via two different-length paths still resolves correctly (topological ordering)."""
        # TOP -> MID -> BOTTOM (existing fixture), plus TOP -> BOTTOM directly (shortcut)
        BomItem.objects.create(part=self.top, sub_part=self.bottom, quantity=1)

        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        requirements = shortfall.calculate_shortfall(
            output.pk, hide_no_shortfall=False
        )

        # Both paths ultimately trace back to the same TOP shortfall date
        self.assertEqual(requirements[self.top.pk]['shortfall_date'], IN_5_DAYS)
        self.assertEqual(requirements[self.mid.pk]['shortfall_date'], IN_5_DAYS)
        self.assertEqual(requirements[self.bottom.pk]['shortfall_date'], IN_5_DAYS)


class DatedEventCollectionTests(ShortfallDateFixtureTestCase):
    """Tests for the dated demand/supply event collector functions directly."""

    def test_undated_sales_order_treated_as_immediate(self):
        """A sales order line with no target_date (and no order target_date) resolves to today."""
        self.create_sales_order(self.bottom, 5)

        events = shortfall.get_dated_demand_events([self.bottom.pk])
        self.assertEqual(events[self.bottom.pk], [(TODAY, Decimal(-5))])

    def test_line_target_date_takes_precedence_over_order_target_date(self):
        """A line item's own target_date wins over the order's target_date when both are set."""
        so = SalesOrder.objects.create(
            customer=self.customer, reference='SO-PREC', target_date=IN_20_DAYS
        )
        SalesOrderLineItem.objects.create(
            order=so, part=self.bottom, quantity=5, target_date=IN_5_DAYS
        )

        events = shortfall.get_dated_demand_events([self.bottom.pk])
        self.assertEqual(events[self.bottom.pk], [(IN_5_DAYS, Decimal(-5))])

    def test_purchase_order_pack_quantity_applied(self):
        """Dated supply events apply the supplier part's pack_quantity, like `Part.on_order` does."""
        packed_supplier_part = SupplierPart.objects.create(
            part=self.bottom,
            supplier=self.supplier,
            SKU='SKU-PACK',
            pack_quantity='10',
        )
        po = PurchaseOrder.objects.create(
            supplier=self.supplier, reference='PO-PACK', target_date=IN_5_DAYS
        )
        PurchaseOrderLineItem.objects.create(
            order=po, part=packed_supplier_part, quantity=3
        )

        events = shortfall.get_dated_supply_events([self.bottom.pk])
        self.assertEqual(events[self.bottom.pk], [(IN_5_DAYS, Decimal(30))])

    def test_build_output_is_a_dated_supply_event(self):
        """An active build order's remaining output is a dated supply event for the built part."""
        Build.objects.create(
            part=self.top,
            quantity=10,
            completed=4,
            reference='BO-0001',
            target_date=IN_10_DAYS,
        )

        events = shortfall.get_dated_supply_events([self.top.pk])
        self.assertEqual(events[self.top.pk], [(IN_10_DAYS, Decimal(6))])


class ShortfallDateReportColumnTests(ShortfallDateFixtureTestCase):
    """Verify the 'Shortfall Date' column appears correctly in the generated XLSX."""

    def test_shortfall_date_column_present_and_correct(self):
        """The exported XLSX includes a 'Shortfall Date' column with the correct value."""
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        shortfall.calculate_shortfall(output.pk, hide_no_shortfall=False)

        output.refresh_from_db()
        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn('Shortfall Date', header)

        date_col = header.index('Shortfall Date')
        rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}

        shortfall_date = rows['DateTop'][date_col]
        # openpyxl round-trips date cells as datetime.datetime
        self.assertEqual(shortfall_date.date(), IN_5_DAYS)

    def test_no_shortfall_date_is_blank_in_report(self):
        """A part with no projected shortfall date has a blank 'Shortfall Date' cell."""
        location = StockLocation.objects.create(name='Location')
        StockItem.objects.create(part=self.top, quantity=1000, location=location)
        self.create_sales_order(self.top, 10, target_date=IN_5_DAYS)

        output = self.make_output()
        shortfall.calculate_shortfall(output.pk, hide_no_shortfall=False)

        output.refresh_from_db()
        from openpyxl import load_workbook

        wb = load_workbook(filename=output.output.path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        date_col = header.index('Shortfall Date')
        rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}

        self.assertIsNone(rows['DateTop'][date_col])
