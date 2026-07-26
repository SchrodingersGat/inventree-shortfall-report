"""Functions for determining component shortfall.

Process Goals:

- Determine the overall "requirements" - based on outstanding Sales Orders
- Iterate downward through the BOMs for each top-level part, to determine the requirements for each sub-component
- Aggregate the "total" requirements for each component (based on the requirements of all parent assemblies)
- Determine the shortfall for each component, based on the available stock, on-order quantity and in-production quantity

"""

from __future__ import annotations

import io
import os
from collections import defaultdict
from datetime import date
from decimal import Decimal

import common.models as common_models
import part.models as part_models
import structlog
from dateutil.relativedelta import relativedelta
from django.core.files.base import ContentFile
from django.db.models import DecimalField, F, Q, Sum
from django.db.models.functions import Coalesce
from InvenTree.helpers import current_date, current_time, normalize
from InvenTree.helpers_model import construct_absolute_url
from openpyxl import Workbook
from openpyxl.styles import Font

logger = structlog.get_logger("inventree.shortfall_report")


def get_plugin():
    """Return the plugin instance for this plugin."""

    from plugin import registry

    return registry.get_plugin("component-shortfall")


def update_part_requirements(
    part, required_qty: Decimal, component_data: dict
) -> Decimal:
    """Return requirements for the given part.

    Arguments:
        part: The part to process
        required_qty: The additional quantity required for the part
        component_data: A dict of part requirements (may be updated)

    Returns:
        The *additional* shortfall for this part (not cumulative)
    """

    requirements = component_data.get(part.pk, None) or {}

    # Store the part information against the part
    requirements["part"] = part

    # Fetch (or calculate) the various stock values for this part
    if "stock" not in requirements:
        stock_items = part.stock_entries(in_stock=True, include_variants=False)

        # TODO: Extend filtering of this query, e.g. exclude certain locations, or stock items with certain parameters?

        # Calculate total and "external" stock quantities in a single query
        result = stock_items.aggregate(
            total=Coalesce(Sum("quantity", output_field=DecimalField()), Decimal(0)),
            external=Coalesce(
                Sum(
                    "quantity",
                    filter=Q(location__external=True),
                    output_field=DecimalField(),
                ),
                Decimal(0),
            ),
        )

        requirements["stock"] = result["total"]
        requirements["external_stock"] = result["external"]

    # TODO: What about BOM items which allow variants???
    # TODO: What about BOM substitutes?

    # Calculate the total "on order" quantity for this part
    if "on_order" not in requirements:
        requirements["on_order"] = part.on_order

    # Calculate the total "in production" quantity for this part
    if "in_production" not in requirements:
        requirements["in_production"] = part.quantity_being_built

    # Add in the additional requirements
    requirements["required"] = requirements.get("required", Decimal(0)) + Decimal(
        required_qty
    )

    # TODO: Support offset for "in production" quantity

    initial_shortfall = requirements.get("shortfall", Decimal(0))

    # Calculate the "shortfall" for this part
    requirements["shortfall"] = max(
        0,
        requirements["required"]
        - requirements["stock"]
        - requirements["on_order"]
        - requirements["in_production"],
    )

    # Update the global dict of component data
    component_data[part.pk] = requirements

    # Return the additional shortfall for this part
    return requirements["shortfall"] - initial_shortfall


def get_outstanding_sales_order_parts(
    horizon_date: date | None = None,
) -> dict:
    """Return a dict of outstanding parts (based on open sales orders).

    Returns a dict of part requirements, with the part ID as the key.

    Each element in the dict has the follow values:
    - part: The part object
    - required: The required quantity of the part (for sales order and build orders)

    Arguments:
        - category: Optional category to filter the parts by
        - horizon_date: Optional cutoff date; orders with a target date beyond this are excluded
    """

    from django.db.models import Q
    from order.models import SalesOrderLineItem
    from order.status_codes import SalesOrderStatus, SalesOrderStatusGroups

    # Find all open sales order line items which are not completed
    sales_order_lines = SalesOrderLineItem.objects.filter(
        order__status__in=SalesOrderStatusGroups.OPEN,
        part__virtual=False,
        shipped__lt=F("quantity"),
    ).prefetch_related(
        "part",
    )

    # Optionally exclude pending sales orders
    plugin = get_plugin()
    exclude_pending_sales = plugin.get_setting("EXCLUDE_PENDING_SALES_ORDERS")

    if exclude_pending_sales:
        sales_order_lines = sales_order_lines.exclude(
            order__status=SalesOrderStatus.PENDING
        )

    # Exclude orders whose target date lies beyond the horizon (undated orders are always included)
    if horizon_date:
        sales_order_lines = sales_order_lines.filter(
            Q(order__target_date__isnull=True) | Q(order__target_date__lte=horizon_date)
        )

    outstanding_parts = {}

    for line in sales_order_lines:
        # The queryset already filters shipped__lt=quantity, so this is always > 0
        deficit = line.quantity - line.shipped

        part_data = outstanding_parts.get(line.part.pk, None) or {
            "part": line.part,
            "required": Decimal(0),
        }
        part_data["required"] += deficit
        outstanding_parts[line.part.pk] = part_data

    return outstanding_parts


def get_outstanding_build_order_parts(
    horizon_date: date | None = None,
) -> dict:
    """Return a dict of outstanding parts (based on open build orders).

    Returns a dict of part requirements, with the part ID as the key.

    Each element in the dict has the follow values:
    - part: The part object
    - required: The required quantity of the part (for sales order and build orders)

    Arguments:
        - category: Optional category to filter the parts by
        - horizon_date: Optional cutoff date; build orders with a target date beyond this are excluded
    """

    from build.models import BuildLine
    from build.status_codes import BuildStatus, BuildStatusGroups
    from django.db.models import Q

    # Find all open build order line items which are not completed
    # Here we are interested in the "deficit" quantity for each line item
    # i.e. the quantity which is still required to complete the build order
    # We must take into account the quantity already consumed against this line item
    build_order_lines = BuildLine.objects.filter(
        build__status__in=BuildStatusGroups.ACTIVE_CODES,
        build__part__virtual=False,
        bom_item__sub_part__virtual=False,
        consumed__lt=F("quantity"),
    ).prefetch_related(
        "bom_item__sub_part",
    )

    # Optionally exclude pending build orders
    plugin = get_plugin()
    exclude_pending_builds = plugin.get_setting("EXCLUDE_PENDING_BUILD_ORDERS")

    if exclude_pending_builds:
        build_order_lines = build_order_lines.exclude(build__status=BuildStatus.PENDING)

    # Exclude build orders whose target date lies beyond the horizon (undated builds are always included)
    if horizon_date:
        build_order_lines = build_order_lines.filter(
            Q(build__target_date__isnull=True) | Q(build__target_date__lte=horizon_date)
        )

    outstanding_parts = {}

    for line in build_order_lines:
        # The queryset already filters consumed__lt=quantity, so this is always > 0
        deficit = line.quantity - line.consumed

        part = line.bom_item.sub_part

        part_data = outstanding_parts.get(part.pk, None) or {
            "part": part,
            "required": Decimal(0),
        }
        part_data["required"] += deficit
        outstanding_parts[part.pk] = part_data

    return outstanding_parts


def get_outstanding_parts(
    horizon_date: date | None = None,
    include_build_orders: bool = True,
    include_sales_orders: bool = True,
) -> dict:
    """Return a dict of outstanding parts (based on open sales orders and build orders).

    Arguments:
        horizon_date: Optional cutoff date; orders with a target date beyond this are excluded
        include_build_orders: Whether to include build orders in the calculation (default: True)
        include_sales_orders: Whether to include sales orders in the calculation (default: True)
    """

    # Start with the outstanding sales order parts
    outstanding_parts = {}

    def add_part_info(parts):
        for part_id, part_data in parts.items():
            if part_id in outstanding_parts:
                outstanding_parts[part_id]["required"] += part_data["required"]
            else:
                outstanding_parts[part_id] = part_data

    if include_build_orders:
        bo_parts = get_outstanding_build_order_parts(horizon_date=horizon_date)
        add_part_info(bo_parts)

    if include_sales_orders:
        so_parts = get_outstanding_sales_order_parts(horizon_date=horizon_date)
        add_part_info(so_parts)

    return outstanding_parts


def record_shortfall_parameters(requirements: dict, parameter_template_id: int) -> None:
    """Record shortfall values against parts using the given parameter template.

    For parts with non-zero shortfall, creates or updates the parameter value.
    For parts with zero shortfall, removes any existing parameter value.
    """

    try:
        template = common_models.ParameterTemplate.objects.get(pk=parameter_template_id)
    except (ValueError, common_models.ParameterTemplate.DoesNotExist):
        logger.warning(
            f"record_shortfall_parameters: ParameterTemplate with ID {parameter_template_id} does not exist"
        )
        return

    content_type = part_models.Part.get_content_type()

    shortfall_ids = set()

    to_create = []
    to_update = []

    now = current_time()

    for data in requirements.values():
        part = data["part"]

        shortfall = data.get("shortfall", Decimal(0))

        if shortfall <= 0:
            continue

        shortfall = str(normalize(data.get("shortfall", Decimal(0))))

        shortfall_ids.add(part.pk)

        # Check if the parameter already exists for this part
        if parameter := part.parameters_list.filter(template=template).first():
            parameter.data = shortfall
            to_update.append(parameter)
        else:
            to_create.append(
                common_models.Parameter(
                    model_type=content_type,
                    model_id=part.pk,
                    template=template,
                    data=shortfall,
                    updated=now,
                )
            )

    if to_create:
        print(f"Creating new shortfall parameters for {len(to_create)} parts")
        common_models.Parameter.objects.bulk_create(to_create, batch_size=250)

    if to_update:
        print(f"Updating shortfall parameters for {len(to_update)} parts")
        common_models.Parameter.objects.bulk_update(to_update, ["data"], batch_size=250)

    # Remove any "shortfall" parameters for parts which are no longer in shortfall
    excluded_pks = list(shortfall_ids)

    to_delete = common_models.Parameter.objects.filter(
        template=template,
        model_type=content_type,
    ).exclude(model_id__in=excluded_pks)

    if to_delete.exists():
        print(
            f"Deleting {to_delete.count()} shortfall parameters for parts no longer in shortfall"
        )
        to_delete.delete()


def _resolve_event_date(*candidates: date | None) -> date:
    """Return the first non-None date from the given candidates.

    Falls back to today's date if every candidate is None - undated demand/supply
    is treated as immediate, per the same convention as the rest of the plugin
    (e.g. undated orders are never excluded by the horizon-date filter either).
    """

    for candidate in candidates:
        if candidate is not None:
            return candidate

    return current_date()


def find_shortfall_date(
    current_balance: Decimal, dated_deltas: list[tuple[date, Decimal]]
) -> date | None:
    """Determine the first date at which a running stock balance goes negative.

    Arguments:
        current_balance: The stock balance on hand today (not including any
            on-order/in-production/outstanding-demand quantity)
        dated_deltas: A list of (date, quantity) tuples - positive for incoming
            supply, negative for outgoing demand

    Returns:
        - Today's date, if `current_balance` is already negative
        - The date of the first event which pushes the running balance negative
        - None, if the balance is never projected to go negative
    """

    if current_balance < 0:
        return current_date()

    # Group deltas by date first, so that same-day events are always applied
    # together - otherwise the arbitrary order two same-day events are processed
    # in could incorrectly report a shortfall date where the net effect for that
    # day is actually non-negative.
    delta_by_date = defaultdict(Decimal)

    for event_date, delta in dated_deltas:
        delta_by_date[event_date] += delta

    balance = current_balance

    for event_date in sorted(delta_by_date):
        balance += delta_by_date[event_date]

        if balance < 0:
            return event_date

    return None


def get_dated_demand_events(part_ids) -> dict[int, list[tuple[date, Decimal]]]:
    """Return per-part lists of dated (date, -quantity) demand events.

    Covers direct demand only (this part's own outstanding sales order lines,
    and this part's own outstanding build-order consumption) - BOM-cascaded
    demand from parent assemblies is layered on separately, since it depends on
    the parent's own resolved shortfall date (see `compute_shortfall_dates`).
    """

    from build.models import BuildLine
    from build.status_codes import BuildStatusGroups
    from order.models import SalesOrderLineItem
    from order.status_codes import SalesOrderStatusGroups

    events = defaultdict(list)

    sales_order_lines = SalesOrderLineItem.objects.filter(
        order__status__in=SalesOrderStatusGroups.OPEN,
        part__virtual=False,
        shipped__lt=F("quantity"),
        part_id__in=part_ids,
    ).select_related("order")

    for line in sales_order_lines:
        deficit = line.quantity - line.shipped

        if deficit <= 0:
            continue

        event_date = _resolve_event_date(line.target_date, line.order.target_date)
        events[line.part_id].append((event_date, -deficit))

    build_lines = BuildLine.objects.filter(
        build__status__in=BuildStatusGroups.ACTIVE_CODES,
        build__part__virtual=False,
        bom_item__sub_part__virtual=False,
        consumed__lt=F("quantity"),
        bom_item__sub_part_id__in=part_ids,
    ).select_related("build", "bom_item")

    for line in build_lines:
        deficit = line.quantity - line.consumed

        if deficit <= 0:
            continue

        event_date = _resolve_event_date(line.build.target_date)
        events[line.bom_item.sub_part_id].append((event_date, -deficit))

    return events


def get_dated_supply_events(part_ids) -> dict[int, list[tuple[date, Decimal]]]:
    """Return per-part lists of dated (date, +quantity) supply events.

    Covers direct supply only: this part's own outstanding purchase order
    lines, and outputs of this part's own active build orders (if it is itself
    an assembly being built).
    """

    from build.models import Build
    from build.status_codes import BuildStatusGroups
    from order.models import PurchaseOrderLineItem
    from order.status_codes import PurchaseOrderStatusGroups

    events = defaultdict(list)

    purchase_order_lines = PurchaseOrderLineItem.objects.filter(
        order__status__in=PurchaseOrderStatusGroups.OPEN,
        part__part_id__in=part_ids,
        quantity__gt=F("received"),
    ).select_related("order", "part")

    for line in purchase_order_lines:
        remaining = line.quantity - line.received

        if remaining <= 0:
            continue

        quantity = line.part.base_quantity(remaining)
        event_date = _resolve_event_date(line.target_date, line.order.target_date)
        events[line.part.part_id].append((event_date, quantity))

    builds = Build.objects.filter(
        status__in=BuildStatusGroups.ACTIVE_CODES,
        part_id__in=part_ids,
    )

    for build in builds:
        remaining = build.remaining

        if remaining <= 0:
            continue

        event_date = _resolve_event_date(build.target_date)
        events[build.part_id].append((event_date, remaining))

    return events


def compute_shortfall_dates(requirements: dict, cascade_edges: dict) -> dict:
    """Compute the shortfall date for every part in `requirements`.

    BOM-cascaded demand must be dated using the *parent's own* resolved
    shortfall date (not "now") - so parents have to be fully resolved before
    their children can be. `cascade_edges` (child part ID -> {parent part ID:
    quantity contributed by that parent}) describes this dependency graph;
    parts are processed in topological order (Kahn's algorithm) so that a
    child is only resolved once every contributing parent already has a
    result, regardless of how the BOM graph's path lengths vary.

    Arguments:
        requirements: The populated requirements dict from `calculate_shortfall`
        cascade_edges: dict of {child_part_id: {parent_part_id: quantity}}

    Returns:
        A dict of {part_id: date | None} - None if that part's stock is never
        projected to go negative given its currently known demand/supply.
    """

    part_ids = list(requirements.keys())

    demand_events = get_dated_demand_events(part_ids)
    supply_events = get_dated_supply_events(part_ids)

    in_degree = {pk: len(cascade_edges.get(pk, {})) for pk in part_ids}

    children_of = defaultdict(list)

    for child_pk, parents in cascade_edges.items():
        for parent_pk in parents:
            children_of[parent_pk].append(child_pk)

    shortfall_dates = {}
    queue = [pk for pk in part_ids if in_degree.get(pk, 0) == 0]

    while queue:
        pk = queue.pop(0)

        events = list(demand_events.get(pk, []))

        # Layer in cascaded demand, dated using each contributing parent's own
        # resolved shortfall date - a parent with no projected shortfall date
        # contributes no cascaded demand event at all.
        for parent_pk, quantity in cascade_edges.get(pk, {}).items():
            parent_date = shortfall_dates.get(parent_pk)

            if parent_date is not None:
                events.append((parent_date, -quantity))

        events += supply_events.get(pk, [])

        current_stock = requirements[pk]["stock"]
        shortfall_dates[pk] = find_shortfall_date(current_stock, events)

        for child_pk in children_of.get(pk, []):
            in_degree[child_pk] -= 1

            if in_degree[child_pk] <= 0:
                queue.append(child_pk)

    return shortfall_dates


def calculate_shortfall(
    output_id: int,
    category_id: int | None = None,
    max_bom_depth: int = 50,
    hide_no_shortfall: bool = True,
    horizon_months: int = 12,
    include_build_orders: bool = True,
    include_sales_orders: bool = True,
    include_supplier_data: bool = False,
    parameter_template_id: int | None = None,
) -> dict:
    """Calculate the component shortfall for a given list of component IDs.

    Arguments:
        output_id: The ID of the DataOutput (where to save the results)
        category_id: The ID of the category to filter parts by (optional)
        max_bom_depth: The maximum depth to traverse the BOM when calculating shortfall (default: 50)
        hide_no_shortfall: Whether to hide parts with no shortfall in the report (default: True)
        horizon_months: Only consider orders due within this many months; 0 means no limit (default: 12)
        include_build_orders: Whether to include build orders in the calculation (default: True)
        include_sales_orders: Whether to include sales orders in the calculation (default: True)
        include_supplier_data: Whether to include supplier information in the shortfall report (default: False)
        parameter_template_id: Optional ID of a ParameterTemplate to record shortfall values against

    Returns:
        A dict of part requirements, with the part ID as the key.

        Each element in the dict has the follow values:
        - part: The part object
        - required: The required quantity of the part (for sales order and build orders)
        - stock: The current stock on hand for this part
        - on_order: The quantity of this part currently on order
        - shortfall: The calculated shortfall for this part (required - stock - on_order)
        - shortfall_date: The date this part is projected to go into shortfall (None if not projected to)
    """

    logger.info("Generating component shortfall report")

    try:
        data_output = common_models.DataOutput.objects.get(pk=output_id)
    except common_models.DataOutput.DoesNotExist:
        logger.error(
            f"component_shortfall: DataOutput with ID {output_id} does not exist - cannot save results"
        )
        return

    # If a PartCategory ID is provided, attempt to fetch the category - if it does not exist, then we will simply ignore the category filter and include all parts in the report
    # We will use this to generate a list of categories, to filter the output dataset
    categories = None

    if category_id:
        try:
            category = part_models.PartCategory.objects.get(pk=category_id)
            categories = category.get_descendants(include_self=True)
        except (ValueError, part_models.PartCategory.DoesNotExist):
            logger.warning(
                f"component_shortfall: PartCategory with ID {category_id} does not exist - cannot filter parts"
            )

    horizon_date = (
        current_date() + relativedelta(months=horizon_months)
        if horizon_months > 0
        else None
    )

    # First, determine the set of components which are "on order"
    initial_parts = get_outstanding_parts(
        horizon_date=horizon_date,
        include_build_orders=include_build_orders,
        include_sales_orders=include_sales_orders,
    )

    # Let's keep track of all the requirements, top-to-bottom, in a single dict - keyed by part ID
    # key: part ID
    # - part: Part instance
    # - required: Total required quantity for this part (cumulative)
    # - stock: Current stock on hand for this part
    # - on_order: Quantity of this part currently on order
    # - building: Quantity of this part currently being built
    requirements = {}

    # Keep a list of the components still required to process - start with the initial set of outstanding parts
    # Each entry is a tuple of (part, quantity, level)
    components_to_process = []

    # Cache of non-consumable BOM items per assembly part, keyed by part ID.
    # A part reused across multiple branches of the BOM graph (or required both
    # directly and via a parent assembly) would otherwise have its BOM re-queried
    # once per visit - this memoizes it to a single query per distinct assembly.
    bom_items_cache = {}

    # Record each BOM-cascade edge as it's discovered: {child part ID: {parent
    # part ID: total quantity contributed by that parent}}. Used afterwards to
    # determine *when* each part's cascaded demand should be dated - see
    # `compute_shortfall_dates`.
    cascade_edges = defaultdict(dict)

    # Start with the initial set of outstanding parts
    for data in initial_parts.values():
        part = data["part"]
        required_qty = data["required"]

        components_to_process.append((part, required_qty, 0))

    # Update initial conditions for the data output
    data_output.progress = 0
    data_output.total = len(components_to_process)
    data_output.save()

    while components_to_process:
        part, quantity, level = components_to_process.pop(0)
        data_output.progress += 1

        shortfall = update_part_requirements(part, quantity, requirements)

        # Update every 50 iterations
        if data_output.progress % 50 == 0:
            data_output.save()

        if shortfall <= 0:
            # No shortfall for this part - skip processing any sub-components
            continue

        # Prevent deep recursion into the BOM - if we have reached the maximum level, then we will not process any sub-components
        if level >= max_bom_depth:
            continue

        # Is this an assembly?
        if part.assembly:
            if part.pk not in bom_items_cache:
                bom_items_cache[part.pk] = list(
                    part.get_bom_items(include_virtual=False)
                    .filter(consumable=False)
                    .prefetch_related(
                        "sub_part",
                        "sub_part__category",
                    )
                )

            for item in bom_items_cache[part.pk]:
                sub_part = item.sub_part

                # Calculate the quantity multiplier for this sub-part
                required_qty = item.get_required_quantity(shortfall)

                components_to_process.append((sub_part, required_qty, level + 1))
                data_output.total += 1

                cascade_edges[sub_part.pk][part.pk] = (
                    cascade_edges[sub_part.pk].get(part.pk, Decimal(0)) + required_qty
                )

    # Determine *when* each part is projected to go into shortfall (not just
    # whether/how much) - see `compute_shortfall_dates` for the topological
    # cascade-dating approach.
    shortfall_dates = compute_shortfall_dates(requirements, cascade_edges)

    for pk, data in requirements.items():
        data["shortfall_date"] = shortfall_dates.get(pk)

    # Record shortfall values against parts using the configured parameter template
    if parameter_template_id:
        record_shortfall_parameters(requirements, parameter_template_id)

    # Generate the output data file
    wb = Workbook()
    ws = wb.active

    cols = [
        "Part Name",
        "Part IPN",
        "Assembly",
        "Purchaseable",
        "Category",
        "Current Stock",
        "External Stock",
        "On Order",
        "In Production",
        "Required Quantity",
        "Shortfall",
        "Shortfall Date",
        "Units",
    ]

    if include_supplier_data:
        cols.extend(["Suppliers"])

    ws.append(cols)

    hyperlink_font = Font(color="0563C1", underline="single")

    for data in requirements.values():
        part = data["part"]

        # If category filtering is enabled, and this part does not belong to the selected category (or its descendants), then skip this part
        if categories and part.category not in categories:
            continue

        shortfall = data.get("shortfall", Decimal(0))

        if hide_no_shortfall and shortfall <= 0:
            continue

        data = [
            part.name,
            part.IPN,
            part.assembly,
            part.purchaseable,
            part.category.pathstring if part.category else None,
            Decimal(data["stock"]),
            Decimal(data["external_stock"]),
            Decimal(data["on_order"]),
            Decimal(data["in_production"]),
            Decimal(data["required"]),
            Decimal(data["shortfall"]),
            data.get("shortfall_date"),
            part.units,
        ]

        if include_supplier_data:
            suppliers = list(
                part.supplier_parts.all()
                .values_list("supplier__name", flat=True)
                .distinct()
            )
            suppliers = sorted(suppliers)  # Sort supplier names alphabetically
            data.append(", ".join(suppliers))

        ws.append(data)

        # Generate link for the part
        cell = ws.cell(row=ws.max_row, column=1)
        cell.hyperlink = construct_absolute_url(part.get_absolute_url())
        cell.font = hyperlink_font

        # Generate link for the category
        if part.category:
            cell = ws.cell(row=ws.max_row, column=5)
            cell.hyperlink = construct_absolute_url(part.category.get_absolute_url())
            cell.font = hyperlink_font

    buf = io.BytesIO()
    wb.save(buf)
    datafile = buf.getvalue()

    data_output.mark_complete(
        output=ContentFile(datafile, name="shortfall_report.xlsx")
    )

    return requirements


def format_shortfall_report_html(
    requirements: dict, output: common_models.DataOutput, hide_no_shortfall: bool = True
) -> str:
    """Format the shortfall report as a HTML document."""

    from django.template import Context, Template

    file_path = os.path.join(
        os.path.dirname(__file__),
        "templates",
        "component_shortfall",
        "shortfall_email.html",
    )

    with open(file_path, "r") as f:
        template_content = f.read()

    context_data = {}

    # Add download link
    if output and output.output:
        context_data["download_link"] = construct_absolute_url(output.output.url)

    # Add all the requirements entries
    requirements_list = []

    for entry in requirements.values():
        if hide_no_shortfall and entry.get("shortfall", 0) <= 0:
            continue

        requirements_list.append(
            {
                **entry,
                "part_url": construct_absolute_url(entry["part"].get_absolute_url()),
            }
        )

    context_data["requirements"] = requirements_list

    template = Template(template_content)
    context = Context(context_data)

    data = template.render(context)

    return data
